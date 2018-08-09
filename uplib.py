import openpyxl
from sslib import get_spreadsheet


class UpdatesHandler:
    def __init__(self, sheet_path, csv_url):
        self._sheet = sheet_path
        self._csv = csv_url
        self._id_map = {}  # id->row
        self._size = 0
        self._init_sheet()

    def _init_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Schedule'

        labels = ['id', 'date', 'start', 'end', 'name', 'lecturer', 'where', 'about']
        ws.append(labels)

        ss = get_spreadsheet(self._csv)
        i = 2
        for item in ss:
            labels = [item['id'], item['date'], item['start'], item['end'],
                      item['name'], item['lecturer'], item['where'], item['about']]
            ws.append(labels)
            self._id_map[item['id']] = i
            i += 1
        self._size = len(ss)

        wb.save(filename=self._sheet)

    def _reload_sheet(self, ss):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Schedule'

        labels = ['id', 'date', 'start', 'end', 'name', 'lecturer', 'where', 'about']
        ws.append(labels)

        i = 2
        self._id_map = {}
        for item in ss:
            labels = [item['id'], item['date'], item['start'], item['end'],
                      item['name'], item['lecturer'], item['where'], item['about']]
            ws.append(labels)
            self._id_map[item['id']] = i
            i += 1
        self._size = len(ss)

        wb.save(filename=self._sheet)

    @staticmethod
    def _compare_ids(old, new):
        """
        Сравнивает значения списков и возвращает три списка:
        Значение встречающиеся только в old, -||- только в new, пересчение old и new
        """
        l_1, l_2, l_3 = [], [], []
        for item in old:
            if item in new:
                l_3.append(item)
                new.remove(item)
            else:
                l_1.append(item)
        l_2 = new

        return l_1, l_2, l_3

    def get_lecture_by_id(self, l_id, spreadsheet=None, worksheet=None):
        if spreadsheet is not None:
            for item in spreadsheet:
                if item['id'] == l_id:
                    return item
        else:
            if worksheet is None:
                wb = openpyxl.load_workbook(filename=self._sheet)
                ws = wb['Schedule']
            else:
                ws = worksheet

            row = str(self._id_map[l_id])
            return {
                'id': ws['A' + row].value,
                'date': ws['B' + row].value,
                'start': ws['C' + row].value,
                'end': ws['D' + row].value,
                'name': ws['E' + row].value,
                'lecturer': ws['F' + row].value,
                'where': ws['G' + row].value,
                'about': ws['H' + row].value
            }

    def get_lectures_by_ids(self, lids, worksheet=None):
        if worksheet is None:
            wb = openpyxl.load_workbook(filename=self._sheet)
            ws = wb['Schedule']
        else:
            ws = worksheet
        result = []
        for lid in lids:
            row = str(self._id_map[lid])
            result.append({
                'id': ws['A' + row].value,
                'date': ws['B' + row].value,
                'start': ws['C' + row].value,
                'end': ws['D' + row].value,
                'name': ws['E' + row].value,
                'lecturer': ws['F' + row].value,
                'where': ws['G' + row].value,
                'about': ws['H' + row].value
            })
        return result

    def _decline_lecture(self, l_id, worksheet):
        info = self.get_lecture_by_id(l_id, worksheet=worksheet)
        text = 'Внимание!\nДоклад "{0}" в {1} отменен.'.format(info['name'], info['start'])
        return text

    def _add_lecture(self, l_id, spreadsheet):
        info = self.get_lecture_by_id(l_id, spreadsheet=spreadsheet)
        text = 'Внимание!\nВ расписание добавлен новый доклад "{0}", который начнется в {1}. ' \
               'Докладчик: {2}.'.format(info['name'], info['start'], info['lecturer'])
        return text

    def get_updates(self):
        """
        Возвращает готовый текст сообщения для дальнейшей его отправки ботом
        """
        ss = get_spreadsheet(self._csv)
        wb = openpyxl.load_workbook(filename=self._sheet)
        ws = wb['Schedule']
        result = []

        old_ids = self._id_map.keys()
        new_ids = [item['id'] for item in ss]
        l_1, l_2, l_3 = self._compare_ids(old_ids, new_ids)

        for i in l_1:
            result.append(self._decline_lecture(i, ws))

        for i in l_2:
            result.append(self._add_lecture(i, ss))

        for item in ss:
            flag = 0
            if item['id'] not in l_3:
                continue
            old_item = self.get_lecture_by_id(item['id'], worksheet=ws)
            if item['start'] != old_item['start']:
                flag = 1
            if item['lecturer'] != old_item['lecturer']:
                if flag == 0:
                    flag = 2
                else:
                    flag = 3

            text = 'Внимание!\n'
            if flag == 0:
                continue
            if flag == 1:
                text += 'Время начала доклада "{0}" изменилось.\n' \
                        'Доклад начнется в {1} и закончится в {2}.'.format(item['name'], item['start'], item['end'])
            if flag == 2:
                text += 'Доклад "{0}" будет читать другой лектор.\n' \
                        'Доклад прочтет {1}'.format(item['name'], item['lecturer'])
            if flag == 3:
                text += 'Время начала доклада "{0}" изменилось.\n' \
                        'Доклад начнется в {1} и закончится в {2}.\n' \
                        'Также изменился изменился докладчик.\n' \
                        'Доклад прочтет {3}'.format(item['name'], item['start'], item['end'], item['lecturer'])
            result.append(text)

        self._reload_sheet(ss)

        return result

